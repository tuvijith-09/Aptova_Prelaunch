"""
main.py — Aptova Prelaunch Backend
Endpoints:
  POST /waitlist    → join-the-community.html
  POST /suggest     → suggest.html
  POST /question    → how-it-works.html
  POST /feedback    → feedback form (shown after join)  ← NEW
Run: uvicorn main:app --reload
"""

import os
import io
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, Request, Depends, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import create_engine, Column, Integer, SmallInteger, String, Text, DateTime
from sqlalchemy.orm import sessionmaker, declarative_base
from pydantic import BaseModel, EmailStr
from dotenv import load_dotenv
from urllib.parse import quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ----------------------------------------------------------------
# Load .env
# ----------------------------------------------------------------
load_dotenv()

DB_HOST         = os.getenv("DB_HOST", "localhost")
DB_PORT         = os.getenv("DB_PORT", "3306")
DB_NAME         = os.getenv("DB_NAME", "bizfit_db")
DB_USER         = os.getenv("DB_USER", "root")
DB_PASSWORD     = os.getenv("DB_PASSWORD", "")
DOWNLOAD_SECRET = os.getenv("DOWNLOAD_SECRET", "aptova2026")

DATABASE_URL = (
    f"mysql+pymysql://{DB_USER}:{quote_plus(DB_PASSWORD)}"
    f"@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4&ssl_disabled=true"
)

# ----------------------------------------------------------------
# DB setup
# ----------------------------------------------------------------
engine       = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=1800)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Base         = declarative_base()


# ----------------------------------------------------------------
# Models
# ----------------------------------------------------------------
class Community(Base):
    __tablename__ = "community"
    id         = Column(Integer,      primary_key=True, autoincrement=True)
    full_name  = Column(String(150),  nullable=False)
    email      = Column(String(255),  nullable=False, unique=True)
    phone      = Column(String(20),   nullable=True)
    city       = Column(String(100),  nullable=True)
    ip_address = Column(String(45),   nullable=True)
    created_at = Column(DateTime,     default=datetime.utcnow)


class Suggestion(Base):
    __tablename__ = "suggestions"
    id          = Column(Integer,     primary_key=True, autoincrement=True)
    full_name   = Column(String(150), nullable=False)
    email       = Column(String(255), nullable=False)
    opp_title   = Column(String(255), nullable=False)
    domain      = Column(String(100), nullable=False)
    capital     = Column(String(100), nullable=False)
    description = Column(Text,        nullable=False)
    ip_address  = Column(String(45),  nullable=True)
    created_at  = Column(DateTime,    default=datetime.utcnow)


class Question(Base):
    __tablename__ = "questions"
    id         = Column(Integer,      primary_key=True, autoincrement=True)
    full_name  = Column(String(150),  nullable=False)
    email      = Column(String(255),  nullable=False)
    question   = Column(Text,         nullable=False)
    ip_address = Column(String(45),   nullable=True)
    is_replied = Column(SmallInteger, default=0)
    created_at = Column(DateTime,     default=datetime.utcnow)


# ----------------------------------------------------------------
# NEW: Feedback model
# ----------------------------------------------------------------
class Feedback(Base):
    __tablename__ = "feedback"
    id                      = Column(Integer,     primary_key=True, autoincrement=True)
    email                   = Column(String(255), nullable=False)          # links back to community table
    first_impression        = Column(Text,        nullable=True)           # Q1
    usefulness_rating       = Column(SmallInteger, nullable=True)          # Q2  (1-5)
    improvement_suggestion  = Column(Text,        nullable=True)           # Q3
    willing_to_help         = Column(String(50),  nullable=True)           # Q4  ("Yes", "Maybe", "Not right now")
    contribution_ways       = Column(Text,        nullable=True)           # Q5  comma-separated checkboxes
    ip_address              = Column(String(45),  nullable=True)
    created_at              = Column(DateTime,    default=datetime.utcnow)

class OpportunityDetail(Base):
    __tablename__ = "opportunity_details"
    id              = Column(Integer,  primary_key=True, autoincrement=True)
    opportunity_key = Column(String(255), nullable=False, unique=True)
    long_description = Column(Text, nullable=False)
    created_at      = Column(DateTime, default=datetime.utcnow)
    updated_at      = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ----------------------------------------------------------------
# FastAPI app
# ----------------------------------------------------------------
app = FastAPI(title="Aptova Prelaunch API", version="1.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    print("[OK] Database connected and tables ready.")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ----------------------------------------------------------------
# Schemas
# ----------------------------------------------------------------
class CommunityIn(BaseModel):
    full_name: str
    email: EmailStr
    phone: Optional[str] = None
    city: Optional[str] = None


class SuggestionIn(BaseModel):
    full_name: str
    email: EmailStr
    opp_title: str
    domain: str
    capital: str
    description: str


class QuestionIn(BaseModel):
    full_name: str
    email: EmailStr
    question: str


# ----------------------------------------------------------------
# NEW: Feedback schema
# ----------------------------------------------------------------
class FeedbackIn(BaseModel):
    email: EmailStr
    first_impression: Optional[str] = None
    usefulness_rating: Optional[int] = None          # 1-5
    improvement_suggestion: Optional[str] = None
    willing_to_help: Optional[str] = None
    contribution_ways: Optional[str] = None           # comma-separated string from frontend

class QuestionIn(BaseModel):
    full_name: str
    email: EmailStr
    question: str


class OpportunityDetailIn(BaseModel):
    opportunity_key: str
    long_description: str

# ----------------------------------------------------------------
# ENDPOINT 1: Waitlist
# ----------------------------------------------------------------
@app.post("/waitlist")
def join_community(payload: CommunityIn, request: Request, db=Depends(get_db)):
    existing = db.query(Community).filter(Community.email == payload.email.lower()).first()
    if existing:
        return {"success": False, "message": "You're already on the waitlist!"}

    entry = Community(
        full_name  = payload.full_name.strip(),
        email      = payload.email.lower().strip(),
        phone      = payload.phone,
        city       = payload.city,
        ip_address = request.client.host,
    )
    db.add(entry)
    db.commit()
    return {"success": True, "message": "You're on the list! We'll notify you on launch."}


# ----------------------------------------------------------------
# ENDPOINT 2: Suggestions
# ----------------------------------------------------------------
@app.post("/suggest")
def submit_suggestion(payload: SuggestionIn, request: Request, db=Depends(get_db)):
    entry = Suggestion(
        full_name   = payload.full_name.strip(),
        email       = payload.email.lower().strip(),
        opp_title   = payload.opp_title.strip(),
        domain      = payload.domain.strip(),
        capital     = payload.capital.strip(),
        description = payload.description.strip(),
        ip_address  = request.client.host,
    )
    db.add(entry)
    db.commit()
    return {"success": True, "message": "Thanks! We review every suggestion and add the best ones."}


# ----------------------------------------------------------------
# ENDPOINT 3: Questions
# ----------------------------------------------------------------
@app.post("/question")
def ask_question(payload: QuestionIn, request: Request, db=Depends(get_db)):
    entry = Question(
        full_name  = payload.full_name.strip(),
        email      = payload.email.lower().strip(),
        question   = payload.question.strip(),
        ip_address = request.client.host,
    )
    db.add(entry)
    db.commit()
    return {"success": True, "message": "Got it! We reply to every question within 24 hours."}


# ----------------------------------------------------------------
# ENDPOINT 4 (NEW): Feedback
# ----------------------------------------------------------------
@app.post("/feedback")
def submit_feedback(payload: FeedbackIn, request: Request, db=Depends(get_db)):
    # Validate rating range if provided
    if payload.usefulness_rating is not None:
        if not (1 <= payload.usefulness_rating <= 5):
            return {"success": False, "message": "Rating must be between 1 and 5."}

    entry = Feedback(
        email                  = payload.email.lower().strip(),
        first_impression       = payload.first_impression.strip() if payload.first_impression else None,
        usefulness_rating      = payload.usefulness_rating,
        improvement_suggestion = payload.improvement_suggestion.strip() if payload.improvement_suggestion else None,
        willing_to_help        = payload.willing_to_help.strip() if payload.willing_to_help else None,
        contribution_ways      = payload.contribution_ways.strip() if payload.contribution_ways else None,
        ip_address             = request.client.host,
    )
    db.add(entry)
    db.commit()
    return {"success": True, "message": "Thank you for your feedback! It means a lot to us."}

@app.get("/opportunities/detail")
def get_opportunity_detail(key: str = Query(...), db=Depends(get_db)):
    row = db.query(OpportunityDetail).filter(OpportunityDetail.opportunity_key == key).first()
    if not row:
        return {"success": False, "message": "No details found for this opportunity"}
    return {"success": True, "long_description": row.long_description}

@app.get("/opportunities/detail")
def get_opportunity_detail(key: str = Query(...), db=Depends(get_db)):
    row = db.query(OpportunityDetail).filter(OpportunityDetail.opportunity_key == key).first()
    if not row:
        return {"success": False, "message": "No details found for this opportunity"}
    return {"success": True, "long_description": row.long_description}


@app.post("/opportunities/detail")
def upsert_opportunity_detail(payload: OpportunityDetailIn, db=Depends(get_db)):
    row = db.query(OpportunityDetail).filter(
        OpportunityDetail.opportunity_key == payload.opportunity_key
    ).first()

    if row:
        row.long_description = payload.long_description
        row.updated_at = datetime.utcnow()
    else:
        row = OpportunityDetail(
            opportunity_key=payload.opportunity_key,
            long_description=payload.long_description
        )
        db.add(row)

    db.commit()
    return {"success": True, "message": "Saved"}

# ----------------------------------------------------------------
# Health check
# ----------------------------------------------------------------
@app.get("/health")
def health():
    return {"status": "ok", "service": "aptova Prelaunch API"}


# ----------------------------------------------------------------
# Helper: style Excel workbook
# ----------------------------------------------------------------
def style_excel(wb, ws, headers):
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    for col in ws.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
            cell.border = thin_border
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = 'A2'


# ----------------------------------------------------------------
# DOWNLOAD: Waitlist as Excel
# ----------------------------------------------------------------
@app.get("/download/waitlist")
def download_waitlist(key: str = Query(...), db=Depends(get_db)):
    if key != DOWNLOAD_SECRET:
        return {"error": "Invalid key"}

    rows     = db.query(Community).order_by(Community.created_at.desc()).all()
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Community"
    headers  = ["ID", "Full Name", "Email", "Phone", "City", "IP Address", "Created At"]
    style_excel(wb, ws, headers)

    for row in rows:
        ws.append([
            row.id, row.full_name, row.email,
            row.phone or "", row.city or "", row.ip_address or "",
            str(row.created_at) if row.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"waitlist_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ----------------------------------------------------------------
# DOWNLOAD: Suggestions as Excel
# ----------------------------------------------------------------
@app.get("/download/suggestions")
def download_suggestions(key: str = Query(...), db=Depends(get_db)):
    if key != DOWNLOAD_SECRET:
        return {"error": "Invalid key"}

    rows     = db.query(Suggestion).order_by(Suggestion.created_at.desc()).all()
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Suggestions"
    headers  = ["ID", "Full Name", "Email", "Opportunity Title", "Domain", "Capital", "Description", "IP Address", "Created At"]
    style_excel(wb, ws, headers)

    for row in rows:
        ws.append([
            row.id, row.full_name, row.email,
            row.opp_title, row.domain, row.capital, row.description,
            row.ip_address or "", str(row.created_at) if row.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"suggestions_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ----------------------------------------------------------------
# DOWNLOAD: Questions as Excel
# ----------------------------------------------------------------
@app.get("/download/questions")
def download_questions(key: str = Query(...), db=Depends(get_db)):
    if key != DOWNLOAD_SECRET:
        return {"error": "Invalid key"}

    rows     = db.query(Question).order_by(Question.created_at.desc()).all()
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Questions"
    headers  = ["ID", "Full Name", "Email", "Question", "Replied", "IP Address", "Created At"]
    style_excel(wb, ws, headers)

    for row in rows:
        ws.append([
            row.id, row.full_name, row.email, row.question,
            "Yes" if row.is_replied else "No",
            row.ip_address or "", str(row.created_at) if row.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"questions_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ----------------------------------------------------------------
# DOWNLOAD (NEW): Feedback as Excel
# ----------------------------------------------------------------
@app.get("/download/feedback")
def download_feedback(key: str = Query(...), db=Depends(get_db)):
    if key != DOWNLOAD_SECRET:
        return {"error": "Invalid key"}

    rows     = db.query(Feedback).order_by(Feedback.created_at.desc()).all()
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Feedback"
    headers  = [
        "ID", "Email", "First Impression", "Usefulness Rating (1-5)",
        "Improvement Suggestion", "Willing to Help", "Contribution Ways",
        "IP Address", "Created At"
    ]
    style_excel(wb, ws, headers)

    for row in rows:
        ws.append([
            row.id,
            row.email,
            row.first_impression or "",
            row.usefulness_rating if row.usefulness_rating is not None else "",
            row.improvement_suggestion or "",
            row.willing_to_help or "",
            row.contribution_ways or "",
            row.ip_address or "",
            str(row.created_at) if row.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"feedback_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})